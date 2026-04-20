from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

hdr = Font(bold=True, color='FFFFFF', size=11, name='Arial')
hdr_fill = PatternFill('solid', fgColor='0D4F3D')
cell_font = Font(size=10, name='Arial')
blue_font = Font(size=10, name='Arial', color='0000FF', bold=True)
thin = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)
center = Alignment(horizontal='center', vertical='center')
left_al = Alignment(horizontal='left', vertical='center', wrap_text=True)
note_font = Font(italic=True, color='888888', size=9, name='Arial')

def write_header(ws, headers):
    for col, val in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=val)
        c.font = hdr
        c.fill = hdr_fill
        c.alignment = center
        c.border = thin

def write_rows(ws, data, blue_cols=None, left_cols=None):
    if blue_cols is None:
        blue_cols = []
    if left_cols is None:
        left_cols = []
    for i, row in enumerate(data, 2):
        for j, val in enumerate(row, 1):
            c = ws.cell(row=i, column=j, value=val)
            c.font = blue_font if j in blue_cols else cell_font
            c.border = thin
            c.alignment = left_al if j in left_cols else center

def set_widths(ws, widths):
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

# ===== 1. ÉVÉNEMENTS =====
ws1 = wb.active
ws1.title = 'Evenements'
write_header(ws1, ['ID', 'Label', 'Emoji', 'Actif'])
write_rows(ws1, [
    ('anniversaire', 'Anniversaire', '\U0001F382', 'OUI'),
    ('corporate', 'Corporate', '\U0001F4BC', 'OUI'),
    ('cocktail', 'Cocktail', '\U0001F378', 'OUI'),
    ('prive', 'Priv\u00e9 / Famille', '\U0001F468\u200D\U0001F469\u200D\U0001F467\u200D\U0001F466', 'OUI'),
    ('autre', 'Autre', '\u2728', 'OUI'),
], blue_cols=[2, 4], left_cols=[2])
set_widths(ws1, {'A': 18, 'B': 22, 'C': 10, 'D': 10})
ws1.cell(row=8, column=1, value='Colonne D = OUI/NON pour activer/d\u00e9sactiver').font = note_font
ws1.cell(row=9, column=1, value='Colonne B (bleu) = modifiable librement').font = note_font

# ===== 2. STYLES =====
ws2 = wb.create_sheet('Styles')
write_header(ws2, ['ID', 'Label', 'Emoji', 'Description'])
write_rows(ws2, [
    ('verrines', 'Verrines & Bouch\u00e9es', '\U0001F942', 'Pi\u00e8ces individuelles \u00e9l\u00e9gantes'),
    ('buffet', 'Buffet', '\U0001F37D\uFE0F', 'Grands bols en libre-service'),
    ('table', 'Repas \u00e0 table', '\U0001FAD5', 'Petits bols sur table'),
    ('box', 'Plateau-repas', '\U0001F371', 'Plateaux individuels complets'),
], blue_cols=[2, 4], left_cols=[2, 4])
set_widths(ws2, {'A': 14, 'B': 24, 'C': 10, 'D': 45})

# ===== 3. FORMULES =====
ws3 = wb.create_sheet('Formules & Prix')
write_header(ws3, ['Style', 'Taille', 'Label', 'Prix/pers (\u20ac)', 'Pi\u00e8ces incluses', 'Mezz\u00e9s froids', 'Prot\u00e9ines', 'Bases', 'Boissons incluses', 'Populaire'])
write_rows(ws3, [
    ('verrines', 'S', 'Essentiel', 15, 6, '', '', '', 'NON', 'NON'),
    ('verrines', 'M', 'G\u00e9n\u00e9reux', 22, 9, '', '', '', 'NON', 'OUI'),
    ('verrines', 'L', 'Prestige', 28, 12, '', '', '', 'OUI', 'NON'),
    ('buffet', 'S', 'Essentiel', 18, '', 4, 1, 1, 'NON', 'NON'),
    ('buffet', 'M', 'G\u00e9n\u00e9reux', 25, '', 5, 2, 2, 'NON', 'OUI'),
    ('buffet', 'L', 'Prestige', 35, '', 7, 3, 2, 'OUI', 'NON'),
    ('table', 'S', 'Essentiel', 20, '', 5, 1, 1, 'NON', 'NON'),
    ('table', 'M', 'G\u00e9n\u00e9reux', 28, '', 6, 2, 2, 'NON', 'OUI'),
    ('table', 'L', 'Prestige', 38, '', 8, 3, 2, 'OUI', 'NON'),
], blue_cols=[3, 4, 5, 6, 7, 8, 9, 10])
for col in range(1, 11):
    ws3.column_dimensions[get_column_letter(col)].width = 16

# ===== 4. VERRINES =====
ws4 = wb.create_sheet('Verrines')
write_header(ws4, ['ID', 'Nom', 'Prix (\u20ac)', 'Premium', 'Extra premium (\u20ac)', 'Tags', 'Actif'])
write_rows(ws4, [
    ('vr1', 'Verrine houmous pistache', 2.5, 'NON', '', 'vegan, sg', 'OUI'),
    ('vr2', 'Verrine moutabal', 2.5, 'NON', '', 'vegan, sg', 'OUI'),
    ('vr3', 'Verrine taboul\u00e9', 2.5, 'NON', '', 'vegan', 'OUI'),
    ('vr4', 'Verrine labn\u00e9 grenade', 2.5, 'NON', '', 'vege', 'OUI'),
    ('vr5', 'Verrine fattouche', 3, 'NON', '', 'vegan', 'OUI'),
    ('vr6', 'Mini falafel + tarator', 2.5, 'NON', '', 'vegan', 'OUI'),
    ('vr7', 'Feuilles de vigne (\u00d72)', 2.5, 'NON', '', 'vegan', 'OUI'),
    ('vr8', 'Mini kebb\u00e9', 3, 'OUI', 0.5, 'halal', 'OUI'),
    ('vr9', 'Verrine mouhamara', 3, 'OUI', 0.5, 'vegan', 'OUI'),
    ('vr10', 'Mini taouk brochette', 3.5, 'OUI', 1, 'halal', 'OUI'),
    ('vr11', 'Mini shawarma cup', 3.5, 'OUI', 1, 'halal', 'OUI'),
], blue_cols=[2, 3, 4, 5, 6, 7], left_cols=[2])
set_widths(ws4, {'A': 8, 'B': 30, 'C': 12, 'D': 12, 'E': 18, 'F': 16, 'G': 10})

# ===== 5. MEZZÉS FROIDS =====
ws5 = wb.create_sheet('Mezzes Froids')
write_header(ws5, ['ID', 'Nom', 'Description', 'Prix (\u20ac)', 'Tags', 'Actif'])
write_rows(ws5, [
    ('mf1', 'Houmous', 'Pois chiches, tahini, citron', 3, 'vegan, sg', 'OUI'),
    ('mf2', 'Moutabal', 'Aubergines grill\u00e9es, tahini', 3, 'vegan, sg', 'OUI'),
    ('mf3', 'Labn\u00e9', 'Fromage frais, huile d\'olive, menthe', 3, 'vege', 'OUI'),
    ('mf4', 'Taboul\u00e9 vert', 'Persil, menthe, boulgour, citron', 3, 'vegan', 'OUI'),
    ('mf5', 'Salade fattouche', 'Laitue, pain grill\u00e9, sumac', 3.5, 'vegan', 'OUI'),
    ('mf6', 'Feuilles de vigne', 'Riz, herbes, citron', 3.5, 'vegan', 'OUI'),
    ('mf7', 'Salade de choux', 'Chou blanc, vinaigrette citron', 2.5, 'vegan', 'OUI'),
    ('mf8', 'Toum', 'Cr\u00e8me d\'ail mont\u00e9e', 1.5, 'vegan', 'OUI'),
], blue_cols=[2, 3, 4, 5, 6], left_cols=[2, 3])
set_widths(ws5, {'A': 8, 'B': 22, 'C': 35, 'D': 12, 'E': 16, 'F': 10})

# ===== 6. PROTÉINES =====
ws6 = wb.create_sheet('Proteines')
write_header(ws6, ['ID', 'Nom', 'Description', 'Prix (\u20ac)', 'Tags', 'Actif'])
write_rows(ws6, [
    ('mc1', 'Falafel \u00d74 + tarator', 'Pois chiches, herbes, s\u00e9same', 4, 'vegan', 'OUI'),
    ('mc2', 'Shawarma poulet + toum', 'Poulet marin\u00e9, \u00e9pices, ail', 5, 'halal', 'OUI'),
    ('mc3', 'Taouk poulet + toum', 'Brochette grill\u00e9e, yaourt', 5, 'halal', 'OUI'),
    ('mc6', 'Shawarma boeuf', 'Boeuf effil\u00e9 aux \u00e9pices', 6, 'halal', 'OUI'),
    ('mc7', 'Chou-fleur zaatar', 'Chou-fleur r\u00f4ti, curcuma', 4, 'vegan, sg', 'OUI'),
], blue_cols=[2, 3, 4, 5, 6], left_cols=[2, 3])
set_widths(ws6, {'A': 8, 'B': 26, 'C': 32, 'D': 12, 'E': 16, 'F': 10})

# ===== 7. BASES =====
ws7 = wb.create_sheet('Bases')
write_header(ws7, ['ID', 'Nom', 'Description', 'Prix (\u20ac)', 'Tags', 'Actif'])
write_rows(ws7, [
    ('mc4', 'Riz \u00e0 la vermicelle', 'Riz basmati, vermicelles dor\u00e9s', 3, 'vege', 'OUI'),
    ('mc5', 'Pomme de terre saut\u00e9e', 'Dor\u00e9es aux herbes', 3, 'vegan', 'OUI'),
    ('bs3', 'Salade verte', 'Laitue, concombre, tomates', 2.5, 'vegan, sg', 'OUI'),
    ('bs4', 'Boulgour', 'Boulgour nature', 3, 'vegan', 'OUI'),
], blue_cols=[2, 3, 4, 5, 6], left_cols=[2, 3])
set_widths(ws7, {'A': 8, 'B': 26, 'C': 32, 'D': 12, 'E': 16, 'F': 10})

# ===== 8. PLATEAUX-REPAS =====
ws8 = wb.create_sheet('Plateaux-repas')
write_header(ws8, ['ID', 'Nom', 'Contenu', 'Prix (\u20ac)', 'Tags', 'Actif'])
write_rows(ws8, [
    ('bx1', 'Plateau V\u00e9g\u00e9tarien', 'Houmous + moutabal + falafel \u00d74 + taboul\u00e9 + fattouche + feuilles de vigne + riz + dessert', 14, 'vegan', 'OUI'),
    ('bx2', 'Plateau Taouk Poulet', 'Houmous + taouk grill\u00e9 + riz vermicelle + taboul\u00e9 + toum + dessert', 16, 'halal', 'OUI'),
    ('bx3', 'Plateau Shawarma Poulet', 'Houmous + shawarma poulet + riz vermicelle + fattouche + toum + dessert', 16, 'halal', 'OUI'),
    ('bx4', 'Plateau Shawarma Boeuf', 'Houmous + shawarma boeuf + riz vermicelle + moutabal + toum + dessert', 18, 'halal', 'OUI'),
    ('bx5', 'Plateau Mixte Grillades', 'Houmous + taouk + shawarma poulet + riz + taboul\u00e9 + toum + dessert', 19, 'halal', 'OUI'),
], blue_cols=[2, 3, 4, 5, 6], left_cols=[2, 3])
set_widths(ws8, {'A': 8, 'B': 26, 'C': 75, 'D': 12, 'E': 12, 'F': 10})

# ===== 9. BOISSONS =====
ws9 = wb.create_sheet('Boissons')
write_header(ws9, ['ID', 'Nom', 'Prix/pers (\u20ac)', 'Actif'])
write_rows(ws9, [
    ('bv1', 'Eau plate (50cl)', 2, 'OUI'),
    ('bv2', 'Eau p\u00e9tillante', 2.5, 'OUI'),
    ('bv3', 'Coca / Sprite / Orangina', 2.5, 'OUI'),
    ('bv4', 'Limonade maison', 3.5, 'OUI'),
    ('bv5', 'Jus de rose maison', 3.5, 'OUI'),
    ('bv6', 'Jus de m\u00fbre maison', 3.5, 'OUI'),
    ('bv7', 'Pack boissons complet', 5.5, 'OUI'),
], blue_cols=[2, 3, 4], left_cols=[2])
set_widths(ws9, {'A': 8, 'B': 28, 'C': 14, 'D': 10})

# ===== 10. DESSERTS =====
ws10 = wb.create_sheet('Desserts')
write_header(ws10, ['ID', 'Nom', 'Prix/pers (\u20ac)', 'Actif'])
write_rows(ws10, [
    ('ds1', 'Baklawa (pi\u00e8ce)', 2.5, 'OUI'),
    ('ds2', 'Mouhalabiyeh (flan libanais)', 4, 'OUI'),
    ('ds3', 'Halawa pistache', 3.5, 'OUI'),
    ('ds4', 'Yaourt confiture de rose', 3.5, 'OUI'),
    ('ds5', 'Plateau desserts mixte', 6, 'OUI'),
], blue_cols=[2, 3, 4], left_cols=[2])
set_widths(ws10, {'A': 8, 'B': 32, 'C': 14, 'D': 10})

# ===== 11. CODES PROMO =====
ws11 = wb.create_sheet('Codes Promo')
write_header(ws11, ['Code', 'R\u00e9duction (%)', 'Actif'])
write_rows(ws11, [
    ('libanais', '5%', 'OUI'),
], blue_cols=[1, 2, 3])
set_widths(ws11, {'A': 20, 'B': 16, 'C': 10})
ws11.cell(row=4, column=1, value='Ajoutez des lignes pour cr\u00e9er de nouveaux codes').font = note_font

output = 'C:/1- Marwan/Claude/03_YE/02_web/v4/config-traiteur.xlsx'
wb.save(output)
print(f'OK: {output}')
print(f'Onglets: {wb.sheetnames}')
